"""
Pruebas de ia/reportes_ia.py: generación de reportes IA (PDF, Word,
Excel) a partir de una ejecución persistida en BD. Usa una base de
datos SQLite temporal y aislada: NUNCA toca database/sigpda.db.
"""
import io

import pandas as pd
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from database.modelos import Base, Plato, Restaurante
import ia.persistencia_bd as persistencia_bd
import ia.reportes_ia as reportes_ia


def _resultado_comparacion_falso():
    fechas = pd.date_range("2026-01-01", periods=40, freq="D")
    serie = pd.DataFrame({"fecha": fechas, "cantidad": range(40)})
    df_plato = pd.DataFrame({"id_plato": 1, "fecha": fechas, "cantidad": range(40)})
    return {
        "modelo_ganador": "holt_winters",
        "modelo_ganador_legible": "Holt-Winters",
        "mae_ganador": 2.0,
        "criterio_seleccion": "menor MAE",
        "metricas_por_modelo": {
            "arima": {"mae": 3.0, "rmse": 3.5, "mape": 20.0, "smape": 18.0, "r2": 0.7},
            "prophet": {"mae": 4.0, "rmse": 4.2, "mape": 25.0, "smape": 22.0, "r2": 0.6},
            "holt_winters": {"mae": 2.0, "rmse": 2.5, "mape": 15.0, "smape": 14.0, "r2": 0.8},
            "transformer_random_forest": {"mae": 2.6, "rmse": 3.0, "mape": 18.0, "smape": 17.0, "r2": 0.72},
            "transformer_gradient_boosting": {"mae": 2.5, "rmse": 2.8, "mape": 17.0, "smape": 16.0, "r2": 0.75},
        },
        "info_modelos": {
            "arima": {"orden": "ARIMA(1,1,1)"},
            "holt_winters": {"hiperparametros": {"trend": "add", "damped_trend": True}},
        },
        "_serie": serie,
        "_df_plato": df_plato,
    }


@pytest.fixture()
def sesion_prueba(tmp_path, monkeypatch):
    ruta_db = tmp_path / "test_reportes.db"
    engine = create_engine(f"sqlite:///{ruta_db}")
    Base.metadata.create_all(engine)
    SesionPrueba = sessionmaker(bind=engine)

    monkeypatch.setattr(persistencia_bd, "obtener_sesion", lambda: SesionPrueba())
    monkeypatch.setattr(reportes_ia, "obtener_sesion", lambda: SesionPrueba())

    sesion = SesionPrueba()
    sesion.add(Restaurante(id_restaurante=1, nombre_comercial="Restaurante Prueba"))
    sesion.add(Plato(id_plato=1, id_restaurante=1, nombre="Ceviche de prueba", categoria="Entradas", precio_venta=20))
    sesion.commit()
    sesion.close()

    yield SesionPrueba


def _crear_ejecucion_completa() -> int:
    resultado = _resultado_comparacion_falso()
    pruebas = [
        {"prueba": "friedman", "estadistico": 5.2, "p_valor": 0.02, "significativo": True,
         "interpretacion": "Diferencia significativa."},
        {"prueba": "wilcoxon", "modelo_a": "holt_winters", "modelo_b": "arima",
         "estadistico": 3.0, "p_valor": 0.03, "p_valor_ajustado": 0.06, "significativo": False,
         "interpretacion": "Sin diferencia tras la corrección."},
        {"prueba": "diebold_mariano", "modelo_a": "arima", "modelo_b": "holt_winters",
         "estadistico": -1.2, "p_valor": 0.23, "significativo": False,
         "interpretacion": "Sin diferencia significativa."},
    ]
    folds_por_modelo = {
        "holt_winters": {
            "folds": [{
                "numero_fold": 1, "n_train": 20, "n_val": 5, "mae": 2.1, "rmse": 2.4,
                "mape": 14.0, "smape": 13.5, "r2": 0.79, "tiempo_entrenamiento": 0.02,
                "fecha_inicio_train": "2026-01-01", "fecha_fin_train": "2026-01-20",
                "fecha_inicio_val": "2026-01-21", "fecha_fin_val": "2026-01-25",
            }],
        },
    }
    return persistencia_bd.guardar_ejecucion_comparacion(
        id_plato=1, id_usuario=1, resultado_comparacion=resultado,
        dias_adelante=7, clima=2, evento=0, duracion_segundos=12.3,
        pruebas_estadisticas=pruebas, folds_por_modelo=folds_por_modelo,
        metadata_extra={
            "eda_resumen": {
                "resumen": {"registros_transacciones": 40},
                "estadisticas_descriptivas": {
                    "media": 5.0, "mediana": 5.0, "desviacion_estandar": 1.2,
                    "minimo": 1, "maximo": 10, "q1": 3, "q3": 7,
                },
                "advertencias": ["Dataset con cobertura limitada."],
            },
            "predicciones_futuras": [
                {"fecha": "2026-02-10", "demanda_estimada": 6.5, "recomendacion": 7, "riesgo": "bajo"},
            ],
            "interpretacion": ["El modelo Holt-Winters fue seleccionado por su menor MAE."],
            "hiperparametros_tuning": {
                "holt_winters": {
                    "aplicable": True, "mejor_hiperparametros": {"trend": "add", "damped_trend": True},
                    "mejor_valor": 2.0, "n_combinaciones": 4, "tiempo_total": 0.5, "semilla": 42,
                    "fecha_ejecucion": "2026-02-01T00:00:00",
                },
                "arima": {"aplicable": False, "motivo": "no aplicable en esta prueba"},
            },
            "modelo_ganador_legible": "Holt-Winters",
            "mae_ganador": 2.0,
        },
    )


class TestObtenerDatosEjecucion:
    def test_ejecucion_inexistente_retorna_none(self, sesion_prueba):
        assert reportes_ia.obtener_datos_ejecucion(999) is None

    def test_reconstruye_toda_la_informacion(self, sesion_prueba):
        id_ejecucion = _crear_ejecucion_completa()
        datos = reportes_ia.obtener_datos_ejecucion(id_ejecucion)

        assert datos is not None
        assert datos["nombre_plato"] == "Ceviche de prueba"
        assert datos["modelo_ganador"] == "holt_winters"
        assert len(datos["modelos"]) == 5

        hw = next(m for m in datos["modelos"] if m["modelo"] == "holt_winters")
        assert len(hw["folds"]) == 1
        assert hw["folds"][0]["mae"] == 2.1

        assert len(datos["pruebas_estadisticas"]) == 3
        assert datos["eda_resumen"]["estadisticas_descriptivas"]["media"] == 5.0
        assert datos["predicciones_futuras"][0]["demanda_estimada"] == 6.5
        assert datos["interpretacion"]
        assert datos["hiperparametros_tuning"]["holt_winters"]["mejor_valor"] == 2.0


class TestGenerarPdfIa:
    def test_ejecucion_inexistente_retorna_none(self, sesion_prueba):
        assert reportes_ia.generar_pdf_ia(999) is None

    def test_genera_bytes_pdf_validos(self, sesion_prueba):
        id_ejecucion = _crear_ejecucion_completa()
        contenido = reportes_ia.generar_pdf_ia(id_ejecucion)
        assert contenido is not None
        assert contenido[:4] == b"%PDF"
        assert len(contenido) > 500


class TestGenerarWordIa:
    def test_ejecucion_inexistente_retorna_none(self, sesion_prueba):
        assert reportes_ia.generar_word_ia(999) is None

    def test_genera_bytes_docx_validos(self, sesion_prueba):
        id_ejecucion = _crear_ejecucion_completa()
        contenido = reportes_ia.generar_word_ia(id_ejecucion)
        assert contenido is not None
        assert contenido[:2] == b"PK"  # .docx es un zip (Office Open XML)
        assert len(contenido) > 500


class TestGenerarExcelIa:
    def test_ejecucion_inexistente_retorna_none(self, sesion_prueba):
        assert reportes_ia.generar_excel_ia(999) is None

    def test_genera_bytes_xlsx_validos_con_todas_las_hojas(self, sesion_prueba):
        from openpyxl import load_workbook

        id_ejecucion = _crear_ejecucion_completa()
        contenido = reportes_ia.generar_excel_ia(id_ejecucion)
        assert contenido is not None
        assert contenido[:2] == b"PK"

        wb = load_workbook(io.BytesIO(contenido))
        hojas_esperadas = {
            "Resumen", "EDA", "Modelos", "Validacion cruzada",
            "Hiperparametros", "Pruebas estadisticas", "Predicciones",
        }
        assert hojas_esperadas.issubset(set(wb.sheetnames))

        ws_modelos = wb["Modelos"]
        assert ws_modelos.cell(row=1, column=1).value == "Modelo"
        assert ws_modelos.freeze_panes == "A2"
